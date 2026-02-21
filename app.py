import streamlit as st
import pandas as pd
import io
import re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ================= 1. 网页基础设置 & 究极 UI 美化 =================
st.set_page_config(page_title="教师课时管理系统", page_icon="📚", layout="wide")

st.markdown("""
<style>
    /* 1. 整体背景色微渐变，护眼且高级 */
    .stApp {
        background-color: #f4f7f6;
        background-image: linear-gradient(120deg, #fdfbfb 0%, #ebedee 100%);
        font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
    }
    
    /* 2. 极致压缩顶部空白，让标题置顶 */
    .block-container {
        padding-top: 1.5rem !important;
        padding-bottom: 2rem !important;
        max-width: 96% !important; 
    }

    /* 3. 重新设计高大上的居中主标题 */
    .main-title {
        font-size: 2.2rem;
        font-weight: 800;
        color: #1e3a8a; /* 商务深蓝 */
        text-align: center;
        margin-bottom: 1.5rem;
        padding-bottom: 1rem;
        border-bottom: 1px solid #cbd5e1;
        letter-spacing: 2px;
        text-shadow: 1px 1px 2px rgba(0,0,0,0.05);
    }

    /* 4. 侧边栏美化：纯白背景加浅浅的阴影，制造悬浮感 */
    [data-testid="stSidebar"] {
        background-color: #ffffff;
        box-shadow: 2px 0 12px rgba(0,0,0,0.04);
        border-right: 1px solid #e2e8f0;
    }
    
    /* 5. 导航按钮美化：苹果风圆角胶囊按键，带悬浮动画 */
    div.stButton > button {
        white-space: nowrap !important; 
        font-size: 13px !important;     
        padding: 4px 12px !important;    
        min-height: 32px !important; 
        height: 32px !important;
        width: 100% !important;         
        background-color: #ffffff;      
        color: #4b5563;
        border: 1px solid #d1d5db;
        border-radius: 16px !important; /* 圆角胶囊 */
        box-shadow: 0 1px 2px rgba(0,0,0,0.05);
        transition: all 0.2s ease-in-out; /* 悬浮动画 */
    }
    div.stButton > button:hover {
        background-color: #f0f9ff;
        color: #0284c7;
        border-color: #7dd3fc;
        transform: translateY(-2px); /* 鼠标移上去微微上浮 */
        box-shadow: 0 4px 6px rgba(0,0,0,0.08);
    }
    
    /* 6. 下载按钮的专属尊贵渐变色 */
    div[data-testid="stDownloadButton"] > button {
        background: linear-gradient(to right, #fbbf24, #f59e0b) !important;
        color: white !important;
        border: none !important;
        font-weight: bold;
        letter-spacing: 1px;
        border-radius: 8px !important;
        box-shadow: 0 4px 6px rgba(245, 158, 11, 0.2) !important;
    }
    div[data-testid="stDownloadButton"] > button:hover {
        background: linear-gradient(to right, #f59e0b, #d97706) !important;
        transform: translateY(-2px);
    }

    /* 7. 行标题（分类名）右对齐，与胶囊按钮对齐 */
    .row-title {
        font-size: 14px;
        font-weight: bold;
        color: #475569;
        text-align: right;               
        padding-top: 6px;
        padding-right: 12px;
        white-space: nowrap;
    }
    
    /* 8. 缩小列间距 */
    [data-testid="column"] { padding: 0 5px !important; }
    
    /* 9. 让数据表格变得像白纸一样干净立体 */
    [data-testid="stDataFrame"] {
        border-radius: 8px;
        overflow: hidden;
        box-shadow: 0 2px 8px rgba(0,0,0,0.04);
        border: 1px solid #e5e7eb;
        background-color: #ffffff;
    }
</style>
""", unsafe_allow_html=True)

# 使用 HTML 注入主标题，替代原本默认自带大片空白的 st.title
st.markdown('<div class="main-title">📚 教师排课智能读取与精准统计系统</div>', unsafe_allow_html=True)

if 'all_sheets' not in st.session_state: st.session_state['all_sheets'] = None
if 'current_sheet' not in st.session_state: st.session_state['current_sheet'] = None
if 'global_mode' not in st.session_state: st.session_state['global_mode'] = False

# ================= 新增核心：汇报级 Excel 渲染引擎 =================
def convert_df_to_excel_pro(df, sheet_name, title):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        export_df = df.reset_index()
        export_df.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
        worksheet = writer.sheets[sheet_name]
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        
        max_col = len(export_df.columns)
        max_row = len(export_df) + 3 
        
        cell = worksheet.cell(row=1, column=1, value=title)
        cell.font = Font(size=18, bold=True, color="000000")
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        cell.alignment = center_align
        worksheet.row_dimensions[1].height = 40 
        
        worksheet.row_dimensions[3].height = 25
        for col_idx in range(1, max_col + 1):
            c = worksheet.cell(row=3, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center_align
            c.border = thin_border
            
        for r_idx in range(4, max_row + 1):
            worksheet.row_dimensions[r_idx].height = 20 
            for c_idx in range(1, max_col + 1):
                c = worksheet.cell(row=r_idx, column=c_idx)
                c.alignment = center_align
                c.border = thin_border
                if c_idx == 1: c.font = Font(bold=True)
                    
        for i in range(1, max_col + 1):
            worksheet.column_dimensions[get_column_letter(i)].width = 14 

    return output.getvalue()

# ================= 2. 智能识别与清洗引擎 =================
def clean_excel_data(df):
    is_schedule = False
    for i in range(min(5, len(df))):
        row_str = " ".join(str(x) for x in df.iloc[i].values)
        if "星期" in row_str or re.search(r'\d{4}[-/]\d{2}[-/]\d{2}', row_str):
            is_schedule = True; break
            
    if is_schedule:
        new_cols = []
        for idx, col in enumerate(df.columns):
            c = str(col).strip()
            if pd.isna(col) or c.lower() in ['nan', '', 'unnamed'] or 'unnamed' in c.lower(): c = f"未命名_{idx+1}"
            base = c
            counter = 1
            while c in new_cols: c = f"{base}_{counter}"; counter += 1
            new_cols.append(c)
        df.columns = new_cols
        return df.dropna(how='all', axis=1).dropna(how='all', axis=0)
    else:
        header_idx = -1
        for i in range(min(10, len(df))):
            if any(k in str(df.iloc[i].values) for k in ["姓名", "科目", "类别", "课数"]):
                header_idx = i; break
        if header_idx != -1:
            raw_cols = df.iloc[header_idx].tolist()
            df = df.iloc[header_idx + 1:].reset_index(drop=True)
        else:
            raw_cols = df.columns.tolist() 
            new_cols = []
            for idx, col in enumerate(raw_cols):
                c = str(col).strip()
                if pd.isna(col) or c.lower() in ['nan', '', 'unnamed'] or 'unnamed' in c.lower(): c = f"未命名_{idx+1}"
                base = c
                counter = 1
                while c in new_cols: c = f"{base}_{counter}"; counter += 1
                new_cols.append(c)
            df.columns = new_cols
        return df.dropna(how='all', axis=1).dropna(how='all', axis=0)

# ================= 6. 核心统计算法库 =================
def parse_class_string(val_str):
    val_str = str(val_str).replace(" ", "") 
    ignore = ['0', '0.0', 'nan', 'none', '星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日', '体育', '班会', '国学', '美术', '音乐', '大扫除']
    if not val_str or val_str.lower() in ignore or re.search(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', val_str) or re.search(r'^第[一二三四五六七八九十]+周', val_str):
        return None
        
    count = 1.0
    m_num = re.search(r'(\d+(?:\.\d+)?)$', val_str)
    if m_num:
        if m_num.start() == 0: return None
        count = float(m_num.group(1))
        val_str = val_str[:m_num.start()] 
        
    match = re.match(r'^([\u4e00-\u9fa5a-zA-Z]+?)(高[一二三]|初[一二三]|小[一二三四五六])(.*)$', val_str)
    if match: return {'教师姓名': match.group(1), '课程类别': match.group(2) + match.group(3), '课时数': count}
        
    known_types = ['早自', '正大', '正小', '晚自', '自大', '自小', '辅导', '正课', '早读', '晚修']
    for kt in known_types:
        if val_str.endswith(kt): return {'教师姓名': val_str[:-len(kt)], '课程类别': kt, '课时数': count}
            
    if len(val_str) >= 2: return {'教师姓名': val_str, '课程类别': '常规课', '课时数': count}
    return None

# ================= 3. 侧边栏与全局汇总配置 =================
st.sidebar.header("📁 数据中心")
uploaded_file = st.sidebar.file_uploader("请拖拽或点击上传 Excel (.xlsm/xlsx)", type=["xlsm", "xlsx"])

if uploaded_file is not None and st.session_state['all_sheets'] is None:
    try:
        with st.spinner('正在执行双引擎解析，请稍候...'):
            raw_sheets = pd.read_excel(uploaded_file, sheet_name=None, engine='openpyxl')
            clean_sheets = {}
            for sheet_name, df in raw_sheets.items(): clean_sheets[sheet_name] = clean_excel_data(df)
            st.session_state['all_sheets'] = clean_sheets
            st.session_state['current_sheet'] = list(clean_sheets.keys())[0]
            st.sidebar.success("✅ 文件解析成功！")
    except Exception as e:
        st.error(f"严重错误: {e}")

if st.session_state['all_sheets'] is not None:
    st.sidebar.markdown("---")
    st.sidebar.subheader("🌐 全局统计生成器")
    
    valid_classes = [s for s in st.session_state['all_sheets'].keys() if not any(kw in s for kw in ['总表', '分表', '汇总'])]
    scope = st.sidebar.radio("📌 统计范围选择", ["所有班级 (全校)", "按年级多选", "自定义勾选班级"])
    
    target_classes = []
    if scope == "所有班级 (全校)":
        target_classes = valid_classes
    elif scope == "按年级多选":
        grades = st.sidebar.multiselect("挑选年级", ["高一", "高二", "高三", "一对一"], default=["高三"])
        target_classes = [c for c in valid_classes if any(g in c for g in grades)]
    else:
        target_classes = st.sidebar.multiselect("勾选具体的班级", valid_classes, default=valid_classes[:2])

    st.sidebar.markdown("##### 📍 数据截取设置")
    col_g1, col_g2 = st.sidebar.columns(2)
    with col_g1: g_start_idx = st.number_input("起始列数", min_value=1, value=15)
    with col_g2: g_end_idx = st.number_input("结束列数", min_value=1, value=21)
    
    g_dates = st.sidebar.date_input("🗓️ 限定统计时间段", [])
    
    if st.sidebar.button("🚀 一键生成全局报表", use_container_width=True, type="primary"):
        if len(g_dates) < 1:
            st.sidebar.error("请先选择完整的时间段！")
        elif not target_classes:
            st.sidebar.error("当前没有选定任何班级！")
        else:
            st.session_state['global_mode'] = True
            st.session_state['g_start'] = g_start_idx
            st.session_state['g_end'] = g_end_idx
            st.session_state['g_dates'] = g_dates
            st.session_state['g_targets'] = target_classes
            st.session_state['g_scope'] = scope

# ================= 4. 动态顶部导航 =================
if st.session_state['all_sheets'] is not None:
    all_sheet_names = list(st.session_state['all_sheets'].keys())
    directory_data = {
        "总表 & 汇总": [], "高一年级": [], "高二年级": [], 
        "高三年级": [], "一对一": [], "其他表单": []
    }
    for name in all_sheet_names:
        if "总" in name or "分表" in name or "汇总" in name: directory_data["总表 & 汇总"].append(name)
        elif "高一" in name: directory_data["高一年级"].append(name)
        elif "高二" in name: directory_data["高二年级"].append(name)
        elif "高三" in name: directory_data["高三年级"].append(name)
        elif "一对一" in name: directory_data["一对一"].append(name)
        else: directory_data["其他表单"].append(name)

    # 导航栏顶部不再需要粗分割线，用微间距代替
    st.write("")
    for category, buttons in directory_data.items():
        if not buttons: continue 
        empty_space = 10 - len(buttons) if len(buttons) < 10 else 1
        cols = st.columns([1.2] + [1] * len(buttons) + [empty_space]) 
        with cols[0]:
            st.markdown(f'<div class="row-title">{category} :</div>', unsafe_allow_html=True)
        for i, btn_name in enumerate(buttons):
            with cols[i+1]:
                if st.button(btn_name, key=f"nav_{btn_name}"):
                    st.session_state['current_sheet'] = btn_name
                    st.session_state['global_mode'] = False 
    st.markdown("<hr style='margin: 15px 0px; border: none; border-top: 1px dashed #cbd5e1;'>", unsafe_allow_html=True)

    # ================= 5. 分支判断：全局表 or 单班级表 =================
    if st.session_state['global_mode']:
        g_dates = st.session_state['g_dates']
        f_start = g_dates[0]
        f_end = g_dates[1] if len(g_dates) == 2 else g_dates[0]
        targets = st.session_state['g_targets']
        
        report_title_prefix = "全校" if st.session_state['g_scope'] == "所有班级 (全校)" else "选中班级"
        
        st.markdown(f"### 🌐 【{report_title_prefix}】课时总汇 📅 ({f_start} 至 {f_end})")
        st.info(f"正在扫描以下 {len(targets)} 个班级：{', '.join(targets[:5])}{' ...' if len(targets)>5 else ''}")
        
        all_records = []
        for s_name in targets:
            if s_name not in st.session_state['all_sheets']: continue
            s_df = st.session_state['all_sheets'][s_name]
            
            start_i = max(0, st.session_state['g_start'] - 1)
            end_i = min(len(s_df.columns), st.session_state['g_end'])
            if start_i >= end_i: continue
                
            locked_cols = s_df.columns[start_i:end_i]
            for col in locked_cols:
                current_date = None
                for val in s_df[col]:
                    val_str = str(val).strip()
                    m = re.search(r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})', val_str)
                    if m:
                        try: current_date = pd.to_datetime(m.group(1)).date()
                        except: pass
                        continue
                    
                    if current_date and (f_start <= current_date <= f_end):
                        parsed = parse_class_string(val_str)
                        if parsed:
                            parsed['来源班级'] = s_name
                            parsed['来源日期'] = str(current_date)
                            all_records.append(parsed)
                            
        if all_records:
            stat_df = pd.DataFrame(all_records)
            pivot_df = pd.pivot_table(stat_df, values='课时数', index='教师姓名', columns='课程类别', aggfunc='sum', fill_value=0)
            pivot_df['总计'] = pivot_df.sum(axis=1)
            
            st.success(f"🎉 统计完毕！共 {len(stat_df['教师姓名'].unique())} 位老师上了课，总计 {stat_df['课时数'].sum()} 节。")
            st.dataframe(pivot_df, use_container_width=True)
            
            formal_title = f"【{report_title_prefix}汇总】课时报表 ({f_start} 至 {f_end})"
            excel_data = convert_df_to_excel_pro(pivot_df, sheet_name="数据汇总", title=formal_title)
            st.download_button(
                label=f"⬇️ 导出带商务排版的《{report_title_prefix}汇报表格》",
                data=excel_data, file_name=f"{report_title_prefix}课时报表_{f_start}至{f_end}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            with st.expander("🔍 查看抓取底层明细 (用于排错)"): st.dataframe(stat_df)
        else:
            st.warning("⚠️ 在指定的范围中，未抓取到有效课时！")
            
    else:
        current = st.session_state['current_sheet']
        st.markdown(f"#### 👁️ 当前查看 : 【 {current} 】")
        
        df_current = st.session_state['all_sheets'][current].copy()
        display_df = df_current.astype(str).replace({' 00:00:00': ''}, regex=True).replace({'nan': '', 'None': ''})
        st.dataframe(display_df, use_container_width=True, height=350)

        st.markdown("---")
        tab1, tab2 = st.tabs(["📏 【周课表专用】垂直穿插统计", "📊 【常规明细表】手动选列统计"])
        
        with tab1:
            all_cols = display_df.columns.tolist()
            col_a, col_b = st.columns(2)
            with col_a: start_choice = st.selectbox("🚩 起始列", options=all_cols, index=14 if len(all_cols)>14 else 0)
            with col_b: end_choice = st.selectbox("🏁 结束列", options=all_cols, index=20 if len(all_cols)>20 else len(all_cols)-1)
                
            start_idx, end_idx = all_cols.index(start_choice), all_cols.index(end_choice)
            if start_idx <= end_idx:
                locked_cols = all_cols[start_idx : end_idx + 1]
                all_dates_in_range = set()
                for col in locked_cols:
                    for val in display_df[col]:
                        m = re.search(r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})', str(val).strip())
                        if m:
                            try: all_dates_in_range.add(pd.to_datetime(m.group(1)).date())
                            except: pass
                
                if all_dates_in_range:
                    min_d, max_d = min(all_dates_in_range), max(all_dates_in_range)
                    date_range = st.date_input(f"🗓️ 选择提取区间：", [min_d, max_d])
                    
                    if len(date_range) >= 1:
                        f_start = date_range[0]
                        f_end = date_range[1] if len(date_range) == 2 else date_range[0]
                        
                        if st.button("🚀 开始本班扫描提取", type="primary"):
                            records = []
                            for col in locked_cols:
                                current_date = None
                                for val in display_df[col]:
                                    val_str = str(val).strip()
                                    m = re.search(r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})', val_str)
                                    if m:
                                        try: current_date = pd.to_datetime(m.group(1)).date()
                                        except: pass
                                        continue
                                    
                                    if current_date and (f_start <= current_date <= f_end):
                                        parsed = parse_class_string(val_str)
                                        if parsed: records.append(parsed)
                                            
                            if records:
                                stat_df = pd.DataFrame(records)
                                pivot_df = pd.pivot_table(stat_df, values='课时数', index='教师姓名', columns='课程类别', aggfunc='sum', fill_value=0)
                                pivot_df['总计'] = pivot_df.sum(axis=1)
                                
                                st.success(f"🎉 统计完毕！【{current}】共计 {stat_df['课时数'].sum()} 节课时。")
                                st.dataframe(pivot_df, use_container_width=True)
                                
                                formal_title = f"【{current}】课时统计报表 ({f_start} 至 {f_end})"
                                excel_data = convert_df_to_excel_pro(pivot_df, sheet_name=current, title=formal_title)
                                st.download_button(
                                    label=f"⬇️ 导出带高级排版的《{current}报表》",
                                    data=excel_data, file_name=f"{current}_课时报表_{f_start}至{f_end}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                                with st.expander("🔍 提取明细"): st.dataframe(stat_df)
                            else:
                                st.warning("未找到可识别的课时。")
                else:
                    st.warning("⚠️ 没有扫描到包含日期的行！")

        with tab2:
            available_cols = list(display_df.columns)
            def guess_index(kw):
                for i, c in enumerate(available_cols):
                    if any(k in str(c) for k in kw): return i
                return 0
                
            col1, col2, col3 = st.columns(3)
            with col1: name_col = st.selectbox("👤 【姓名】列", available_cols, index=guess_index(['姓名','教师']))
            with col2: type_col = st.selectbox("🏷️ 【类别】列", available_cols, index=guess_index(['子类','类别']))
            with col3: count_col = st.selectbox("🔢 【数量】列", available_cols, index=guess_index(['课数','课时']))
                
            if st.button("📊 生成常规统计"):
                try:
                    stat_df = df_current.copy()
                    stat_df[count_col] = pd.to_numeric(stat_df[count_col], errors='coerce').fillna(0)
                    stat_df = stat_df[stat_df[name_col].notna()]
                    stat_df = stat_df[stat_df[name_col].astype(str).str.strip() != '']
                    pivot_df = pd.pivot_table(stat_df, values=count_col, index=name_col, columns=type_col, aggfunc='sum', fill_value=0)
                    pivot_df['总计'] = pivot_df.sum(axis=1)
                    st.dataframe(pivot_df, use_container_width=True)
                    
                    formal_title = f"【{current}】常规课时统计"
                    excel_data = convert_df_to_excel_pro(pivot_df, sheet_name=current, title=formal_title)
                    st.download_button(
                        label="⬇️ 导出带高级排版的报表", data=excel_data, file_name=f"{current}_常规课时.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except:
                    st.warning("无法生成，请确认选对了列名！")
else:
    st.info("👆 请先在左侧上传您的 Excel 文件！")